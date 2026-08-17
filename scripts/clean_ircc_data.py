"""
clean_ircc_data.py
-------------------
Transforms the raw IRCC "Permanent Residents by Country of Citizenship"
XLSX (as downloaded by fetch_ircc_data.py) into the long-format CSV the
build script expects: Country, Year, Month, Date, Admissions.

Raw file layout (sheet "PR - CITZ"):
    Row 3   Year headers, one merged block per year (e.g. "2026")
    Row 5   Month headers within each year block: Jan..Dec plus
            "Q1 Total".."Q4 Total" columns (skipped) and a blank
            spacer column between years (skipped)
    Data    One row per country starting after the header rows, until
            the "Total" row
    Values  "--" for suppressed (0-5, redacted), "0" for true zero,
            comma-thousands for everything else (e.g. "12,910")

Usage:
    python clean_ircc_data.py [--input data/raw/EN_ODP-PR-Citz.xlsx] [--output data/clean/pr_citz_long_clean.csv]
"""

import argparse
import os
import sys

import openpyxl
import pandas as pd

SHEET_NAME = "PR - CITZ"
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_NUM = {m: i + 1 for i, m in enumerate(MONTHS)}
DEFAULT_INPUT = os.path.join("data", "raw", "EN_ODP-PR-Citz.xlsx")
DEFAULT_OUTPUT = os.path.join("data", "clean", "pr_citz_long_clean.csv")


def parse_value(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s == "--":
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def find_month_columns(ws):
    """Map each data column to (year, month) using the year row (3) and
    month row (5), forward-filling years across their merged blocks."""
    years = []
    last_year = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if v is not None:
            last_year = v
        years.append(last_year)

    month_cols = []
    for c in range(1, ws.max_column + 1):
        month = ws.cell(row=5, column=c).value
        year = years[c - 1]
        if month in MONTH_NUM and year is not None:
            month_cols.append((c, int(year), month))
    return month_cols


def find_data_rows(ws):
    """Country rows start after the header block and run until the
    'Total' row (or a blank country cell)."""
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Country of Citizenship":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Could not find 'Country of Citizenship' header row")

    start = header_row + 3  # skip the two sub-header rows (quarter, month)
    rows = []
    for r in range(start, ws.max_row + 1):
        country = ws.cell(row=r, column=1).value
        if country is None or str(country).strip() == "":
            break
        if str(country).strip() == "Total":
            break
        rows.append((r, str(country).strip()))
    return rows


def transform(input_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[SHEET_NAME]

    month_cols = find_month_columns(ws)
    data_rows = find_data_rows(ws)

    records = []
    for r, country in data_rows:
        for c, year, month in month_cols:
            admissions = parse_value(ws.cell(row=r, column=c).value)
            records.append({
                "Country": country,
                "Year": year,
                "Month": month,
                "Date": f"{year:04d}-{MONTH_NUM[month]:02d}-01",
                "Admissions": admissions,
            })

    df = pd.DataFrame(records)
    df["Month"] = pd.Categorical(df["Month"], categories=MONTHS, ordered=True)
    df = df.sort_values(["Country", "Year", "Month"]).reset_index(drop=True)
    df["Month"] = df["Month"].astype(str)
    return df


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=DEFAULT_INPUT)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    print(f"Reading:  {args.input}")
    try:
        df = transform(args.input)
    except (FileNotFoundError, ValueError) as e:
        print(f"Clean failed: {e}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    df.to_csv(args.output, index=False)
    print(f"Saved:    {args.output} ({len(df):,} rows, {df['Country'].nunique()} countries)")


if __name__ == "__main__":
    main()
